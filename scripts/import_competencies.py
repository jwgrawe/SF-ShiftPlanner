#!/usr/bin/env python3
"""Import the anonymized competency workbook into data/seed/employee_competencies.csv.

Reads data/source/Kompetanse_Anonymisert.xlsx (requires openpyxl) and writes
one row per (employee, competency) with a status:

  x / X -> qualified
  ?     -> uncertain   (a manager's reminder to assess this competency, D44;
                        NOT eligible for planning)
  anything else -> ignored with a warning (treated as no competency, D44)

Competency columns are mapped to the competency_id values in
data/seed/competency_types.csv (functions map to competencies separately, in
function_competencies.csv). The combined source column
"Sterrad + poliklinikker/løspakk" maps to BOTH split competencies until the
source file gains separate columns (D43/Q20).

Anonymized rows "Employee N" map to employee_id E<NNN>, matching the
source_label column produced by scripts/generate_fake_employees.py.

Usage:  python scripts/import_competencies.py
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE = REPO_ROOT / "data" / "source" / "Kompetanse_Anonymisert.xlsx"
SEED_DIR = REPO_ROOT / "data" / "seed"
OUT = SEED_DIR / "employee_competencies.csv"

STATUS = {"x": "qualified", "?": "uncertain"}


def load_column_mapping() -> dict[str, list[str]]:
    """source_column -> [competency_id, ...] from competency_types.csv."""
    mapping: dict[str, list[str]] = {}
    with open(SEED_DIR / "competency_types.csv", newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            mapping.setdefault(row["source_column"], []).append(row["competency_id"])
    return mapping


def main() -> int:
    if not SOURCE.exists():
        print(f"Source file not found: {SOURCE}")
        return 1

    column_mapping = load_column_mapping()
    ws = load_workbook(SOURCE, data_only=True).worksheets[0]
    headers = [cell.value for cell in ws[2]]

    unknown = [h for h in headers[1:] if h and h not in column_mapping]
    if unknown:
        print(f"Unmapped competency columns (add them to competency_types.csv): {unknown}")
        return 1

    rows: list[tuple[str, str, str]] = []
    warnings: list[str] = []
    n_employees = 0
    for row in ws.iter_rows(min_row=3):
        label = row[0].value
        if label is None:
            continue
        label = str(label).strip()
        if not label.startswith("Employee "):
            print(f"Unexpected row label {label!r} – expected 'Employee N'")
            return 1
        employee_id = f"E{int(label.split()[1]):03d}"
        n_employees += 1
        for cell, header in zip(row[1:], headers[1:]):
            if header is None:
                continue
            mark = str(cell.value).strip().lower() if cell.value is not None else ""
            if not mark:
                continue
            status = STATUS.get(mark)
            if status is None:
                warnings.append(
                    f"{label} / {header!r}: mark {mark!r} ignored (treated as no competency)"
                )
                continue
            for competency_id in column_mapping[header]:
                rows.append((employee_id, competency_id, status))

    rows.sort()
    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["employee_id", "competency_id", "status"])
        writer.writerows(rows)

    for warning in warnings:
        print(f"WARNING: {warning}")
    print(f"Imported {n_employees} employees, {len(rows)} competency rows -> {OUT}")
    per_competency: dict[str, int] = {}
    for _, competency_id, status in rows:
        if status == "qualified":
            per_competency[competency_id] = per_competency.get(competency_id, 0) + 1
    for competency_ids in column_mapping.values():
        for competency_id in competency_ids:
            print(f"  {competency_id:24s} {per_competency.get(competency_id, 0):3d} qualified")
    return 0


if __name__ == "__main__":
    sys.exit(main())
